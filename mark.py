from tkinter import*
from tkinter import ttk,messagebox
from openpyxl import Workbook,load_workbook
import os
root=Tk()
root.geometry("600x500")
root.title("Information collection")

fname="mark.xlsx"
header=["Name","Class and section","Tamil","Englis","Major"]

if not os.path.exists(fname):
    wb=Workbook()
    ws=wb.active
    ws.append(header)
    wb.save(fname)

def clear():
    e1.delete(0,END)
    e2.delete(0,END)
    e3.delete(0,END)
    e4.delete(0,END)
    e5.delete(0,END)


def submit():
    
    data=[e1.get(),e2.get(),e3.get(),e4.get(),e5.get()]
    print(data)
    if all(data):
        if 101 > int(data[2])>0 and 101 > int(data[3])>0 and  101 > int(data[4])>0 :
            wb=load_workbook(fname)
            ws=wb.active
            ws.append(data)
            wb.save(fname)
            messagebox.showinfo("Susscully","your data submitted successfully")
            clear()
        else:
             messagebox.showwarning("Input type error","Please enter correct mark")
             clear()
            
        
    else:
        messagebox.showwarning("Input Error","Pleas fill all required fields")


def search():
    if e6.get()=="":
        wb=load_workbook(fname)
        ws=wb.active
        for i in ws.iter_rows(values_only=True):
            table.insert(parent='',index=END,values=i)
    else:
        wb=load_workbook(fname)
        ws=wb.active
        for i in ws.iter_rows(values_only=True):
            if i[0]==e6.get():
                  if  i[0]=="Name":
                      print("welcome")
                  
            else:
                table.insert(parent='',index=END,values=i)
                
            
    
    

l1=Label(root,text="Enter your name :")
l1.grid(row=0,column=0)
l2=Label(root,text="Enter your Class and section :")
l2.grid(row=1,column=0)
l3=Label(root,text="Enter you Tamil mark :")
l3.grid(row=2,column=0)
l4=Label(root,text="Enter you English Mark :")
l4.grid(row=3,column=0)
l5=Label(root,text="Enter you Major mark :")
l5.grid(row=4,column=0)

e1=Entry(root)
e1.grid(row=0,column=1)
e2=Entry(root)
e2.grid(row=1,column=1)
e3=Entry(root)
e3.grid(row=2,column=1)
e4=Entry(root)
e4.grid(row=3,column=1)
e5=Entry(root)
e5.grid(row=4,column=1)
e6=Entry(root)
e6.grid(row=8,column=0)

Button(root,text="Submit",command=submit).grid(row=5,columnspan=2)
Button(root,text="SEARCH",command=search).grid(row=8,column=1)

table=ttk.Treeview(root,columns=("Name","Class","Tamil","Englis","Major"),show="headings")
table.heading('Name',text="Name")
table.heading('Class',text="class")
table.heading('Tamil',text="Tamil")
table.heading('Englis',text="English")
table.heading('Major',text="Major")
table.grid(row=10,columnspan=5)

         
